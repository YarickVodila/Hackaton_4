import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from datetime import date, timedelta, datetime
from catboost import CatBoostRegressor, Pool,CatBoostClassifier
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from sklearn.metrics import mean_squared_error
from sklearn import metrics
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.models import Sequential
from sklearn import preprocessing
import time
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
import math
#WebPriceId     tDateObserve        tStockStatus        tCurrentPrice


#month_end = int(input('Введите число конца месяца (например 31): '))
#name_file_save = 'test_1.txt'
# f = open('name_file_save', w+)
# f.write(значение)
# f.close()

start_time = time.time()
#WebPriceId     tDateObserve        tStockStatus        tCurrentPrice



chunksize = 10000 #батч
Id_corr = []
file_name = 'DS_train(2020-06--2022-06-21).csv'
array = np.array([]) # Массив для
cor_array = np.array([]) #Массив CPI

id = 1 #Переменная которая содержит id признака которого сейчас рассматриваем
num=1
flag = False
#s = 0
count =0
mean_cpi = 0

#
def data_exel_read(filename_read):

    exel = load_workbook(filename_read) #загружаем ексель файл как объект
    sheet = exel.sheetnames #получаем в виде списка названия листов
    ws = exel[sheet[0]]
    data = ws.values
    Data = pd.DataFrame(data).to_numpy()
    return Data[2][1:]

date_df_y = data_exel_read('Y_train.xlsx')
date_df_y = date_df_y.astype(float)




def processing(array, id):

    global cor_array, mean_cpi,count
    df = pd.DataFrame(data = array, columns= ['WebPriceId','tDateObserve','tStockStatus','tCurrentPrice'])
    array_options = df.to_numpy()
    OutOfStock = 0
    #print(df['tStockStatus'].shape)
    if df['tStockStatus'].value_counts()['InStock'] > 24:
        #dt  = datetime.strptime(array[1][1], '%Y-%m-%d %H:%M:%S.%f').date()
        df = df[df['tStockStatus']=='InStock']
        #df['tCurrentPrice'] = np.where((df.tStockStatus == 'OutOfStock'), True, df.tCurrentPrice)
        array = df.to_numpy()
        #print(array)


        first_inst  = datetime.strptime(array[0][1], '%Y-%m-%d %H:%M:%S.%f').date()
        second_inst = datetime(2022, 5, 31).date() #Важно!!! До какого заполняем



        date_range = pd.Series(pd.date_range(first_inst, periods = (second_inst-first_inst).days+1, freq ='D'))
        #print(array)
        zero_array = np.zeros(date_range.shape[0])

        data_train = pd.DataFrame({'Data': date_range, 'price': zero_array, 'Bool': zero_array})
        data_train['price'] = np.where((data_train.price == 0.0), None, data_train.price)
        data_train['Bool'] = np.where((data_train.Bool == 0.0), None, data_train.Bool)

        munth = pd.to_datetime(data_train.Data[0]).month
        day = pd.to_datetime(data_train.Data[0]).day

        for i in range(array.shape[0]):
            data_train['price'] = np.where((data_train.Data == pd.to_datetime(datetime.strptime(array[i][1], '%Y-%m-%d %H:%M:%S.%f').date())), array[i][3], data_train.price)
            #print(array[i][1])

        data_train = data_train.fillna(method='ffill') #Получаем датафрейм с ценами и датами


        for i in range((datetime(2021, 5, 31).date()-first_inst).days,zero_array.shape[0]): #Не забыть изменить смотря от дня
            if ((pd.to_datetime(data_train.Data[i])+timedelta(days=1)).day == 1): #pd.to_datetime(data_train.Data[i]).day == 20 or
                data_train['Bool'] = np.where((data_train.Data == pd.to_datetime(data_train.Data[i])), True, data_train.Bool)
                #print(pd.to_datetime(data_train.Data[i])+timedelta(days=1))

        data_train.dropna(subset=['Bool'], inplace=True)
        data_train = data_train.drop('Bool', axis=1)
        #print(data_train)
        #cpi = CPI(data_train,id)

        #cor_array = np.array([])
        dataframe_array = data_train.to_numpy()
        data_train['CPI'] = np.zeros(data_train.shape[0])
        #print(data_train)
        #dataframe = dataframe[['price','CPI']].to_numpy()
        #price_first = 0
        #price_second = 0
        CPI = np.array([])

        if ((dataframe_array[0][0]).date()).day == 20: #Не забыть изменить смотря от дня
            dataframe_array = dataframe_array[1:]
            #print(dataframe_array)

        for i in range(0,data_train.shape[0]-1,2):
            #print(data_train.shape[0])
            CPI = np.append(CPI,(((dataframe_array[i+1][1]/dataframe_array[i][1])*100)-100))
        #print(CPI)
            #dataframe['CPI'] = np.where((dataframe.Data == pd.to_datetime((dataframe_array[i+2][0]).date())), ((dataframe_array[i+1][1]/dataframe_array[i][1])*100)-100, dataframe.CPI)

        #print(date_df_y[len(date_df_y)-12:])
        '''
        if CPI.shape[0] == 12 and CPI.sum()!=0 and (np.corrcoef(CPI, date_df_y[len(date_df_y)-CPI.shape[0]:])[0][1]) > 0.86:
            #cor_array = np.append(cor_array,(np.corrcoef(CPI, date_df_y[len(date_df_y)-12:])[0][1]))
            cor_array = np.append(cor_array,(np.array([CPI])))
            print('Корреляция равна = ',np.corrcoef(CPI, date_df_y[len(date_df_y)-CPI.shape[0]:])[0][1], 'id = ', id)
            print(CPI)
            mean_cpi += CPI[-1]
            count+=1
        '''
        if CPI.sum()!=0 and (math.dist(CPI,date_df_y[len(date_df_y)-CPI.shape[0]:])) < 4: #
            #cor_array = np.append(cor_array,(np.corrcoef(CPI, date_df_y[len(date_df_y)-12:])[0][1]))
            cor_array = np.append(cor_array,(np.array([CPI])))
            print('Евклидово расстояние = ',math.dist(CPI,date_df_y[len(date_df_y)-CPI.shape[0]:]), 'id = ', id)
            print(CPI)
            mean_cpi += CPI[-1]
            count+=1

'''
        #return cpi
        #print(data_train.to_string())
        #print(array.shape[0])'''





for chunk in pd.read_csv(file_name, sep='\t', chunksize=chunksize):
    #print(2)
    '''
    if chunk['WebPriceId'].to_numpy()[0] == id:
        array = np.append(array, (chunk.to_numpy()))
    else:
        id = chunk['WebPriceId'].to_numpy()[0] #Обновляем айдишник
        #print(array)
        array = np.array([]) #Очищаем массив
        array = np.append(array, (chunk.to_numpy())) #добавляем первое значение нового айдишника
        print(id)'''
    #print(id in chunk['WebPriceId'].to_numpy())
    #sum+=chunk.to_numpy().shape[0]
    #print(chunk.to_string())
    while id in chunk['WebPriceId'].to_numpy():
        if id-1 in chunk['WebPriceId'].to_numpy() and flag == True:
            #print(np.reshape(array, (int(array.shape[0]/4), 4)))
            array = np.append(array, (chunk[chunk['WebPriceId'] == id-1].to_numpy())) #Добавляем значения
            array = np.reshape(array, (int(array.shape[0]/4), 4))
            # Тут чёт делаем и потом очищаем массив и берём следующий батч
            #cor_array = np.append(cor_array,(processing(array,id)))
            processing(array,id)
            #print(np.reshape(array, (int(array.shape[0]/4), 4)))
            array = np.array([]) #Очищаем массив
            id+=1
            flag = False
            #break

        elif id+1 in chunk['WebPriceId'].to_numpy(): #Если следующее значение есть то обнуляем значение и меняем id
            array = np.append(array, (chunk[chunk['WebPriceId'] == id].to_numpy()))
            # Тут чёт делаем и потом очищаем массив и берём следующий батч
            array = np.reshape(array, (int(array.shape[0]/4), 4))
            #print(pd.DataFrame(data = array, columns= ['WebPriceId','tDateObserve','tStockStatus','tCurrentPrice'])['tStockStatus'].value_counts())
            processing(array,id)
            array = np.array([]) #Очищаем массив
            id+=1
            flag = False
            #break

        elif id+1 not in chunk['WebPriceId'].to_numpy():#Если следующего значения нет
            array = np.append(array, (chunk[chunk['WebPriceId'] == id].to_numpy()))
            id+=1
            flag = True

    #print(id)
    #print(cor_array.shape[0])

    if id >= 15000:
        break
'''
    if cor_array.shape[0] == 48:
        print(cor_array)
        break
'''

name_file_save = 'test_1.txt'
f = open(name_file_save, 'w+')
f.write(str(mean_cpi/count))
f.close()

print(id)
print('Предсказание = :', mean_cpi/count)
# 800 сек рекорд!!! +- 13 мин
print("--- %s Время: ---" % (time.time() - start_time))
